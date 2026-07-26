#!/usr/bin/env python3
"""Run selected GIGASTROKE datasets one at a time with an Excel status ledger."""
from __future__ import annotations

import argparse
import json
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook

HEADERS = ["dataset_id", "ancestry", "outcome", "role", "status", "started_utc",
           "finished_utc", "accepted", "rejected", "message"]


def now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def save(book, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    book.save(temporary); temporary.replace(path)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--config", required=True, type=Path)
    parser.add_argument("--status-xlsx", type=Path)
    parser.add_argument("--force", action="store_true", help="Rerun rows already marked completed.")
    args = parser.parse_args()
    config = json.loads(args.config.read_text(encoding="utf-8"))
    output_dir = Path(config["output_dir"])
    if not output_dir.is_absolute(): output_dir = (args.config.parent / output_dir).resolve()
    status_path = args.status_xlsx or output_dir / "gigastroke_batch_status.xlsx"
    selected = [config["selection"]["eur_discovery"], *config["selection"]["eas_replication_subtypes"]]
    datasets = {item["id"]: item for item in config["datasets"]}

    if status_path.exists():
        book = load_workbook(status_path); sheet = book["status"]
    else:
        book = Workbook(); sheet = book.active; sheet.title = "status"; sheet.append(HEADERS)
    rows = {sheet.cell(row=i, column=1).value: i for i in range(2, sheet.max_row + 1)}
    for dataset_id in selected:
        if dataset_id not in rows:
            item = datasets[dataset_id]
            sheet.append([dataset_id, item["ancestry"], item["outcome"], item["role"], "pending", "", "", "", "", ""])
            rows[dataset_id] = sheet.max_row
    save(book, status_path)

    adapter = Path(__file__).with_name("gigastroke_outcome_adapter.py")
    for dataset_id in selected:
        row = rows[dataset_id]
        if sheet.cell(row, 5).value == "completed" and not args.force: continue
        sheet.cell(row, 5, "running"); sheet.cell(row, 6, now()); sheet.cell(row, 7, "")
        sheet.cell(row, 10, ""); save(book, status_path)
        command = [sys.executable, str(adapter), "--config", str(args.config.resolve()), "--dataset-id", dataset_id]
        result = subprocess.run(command, text=True, stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
        sheet.cell(row, 7, now())
        if result.returncode:
            sheet.cell(row, 5, "failed"); sheet.cell(row, 10, result.stdout[-30000:])
            save(book, status_path)
            raise SystemExit(f"{dataset_id} failed; see {status_path}")
        manifest = json.loads((output_dir / "dataset_manifest.json").read_text(encoding="utf-8"))
        item = next(value for value in manifest if value["dataset_id"] == dataset_id)
        sheet.cell(row, 5, "completed"); sheet.cell(row, 8, item["accepted"]); sheet.cell(row, 9, item["rejected"])
        sheet.cell(row, 10, result.stdout[-30000:]); save(book, status_path)
        print(f"[OK] {dataset_id}: accepted={item['accepted']} rejected={item['rejected']}", flush=True)


if __name__ == "__main__":
    main()
